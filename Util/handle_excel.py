#coding=utf-8
import sys,os
import openpyxl
import configparser

class HandelExcel:

    def getinit(self,section,key):
        cf=configparser.ConfigParser()
        cf.read(r"F:\project\Config\server.ini")
        data=cf.get(section,key)
        return data

    def load_excel(self):
        '''
        加载excle
        '''
        open_excel=openpyxl.load_workbook(r"F:\project\Case\imooc.xlsx")
        return open_excel
    
    def get_sheet_data(self,index=None):
        '''
        加载所有的sheet内容，拿到需要的sheet页
        '''
        sheet_name=self.load_excel().sheetnames
        
        if index==None:
            index=0
        data=self.load_excel()[sheet_name[index]]
        # data=sheet_name[index]
        return data
    
    def get_cell_value(self,row,cols):
        '''
        获取某一个单元各的内容
        '''
        data=self.get_sheet_data().cell(row=row,column=cols).value
        return data
    
    def get_row(self):
        '''
        获取所有的行数
        '''
        row=self.get_sheet_data().max_row
        row-=2
        return row
    def get_row_value(self,row):
        '''
        获取行的值
        '''
        row_list=[]
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_Data(self,row,clos,valus):
        '''写入数据'''
        wb=self.load_excel()
        #激活文件
        wr=wb.active
        #写进文件的位置
        wr.cell(row,clos,valus)
        wb.save(r"F:\project\Case\imooc.xlsx")


excel_data=HandelExcel()

if __name__ == "__main__":
    print(excel_data.getinit("server","base_path"))
    # print(excel_data.get_row_value(4))
    # print(excel_data.excel_write_Data(7,2,"tongguo"))
    # print(excel_data.get_cell_value(2,3))
    # print(rest)

    